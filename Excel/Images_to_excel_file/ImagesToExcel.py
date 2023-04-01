# pip install Pillow
from openpyxl import Workbook # pip install openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
import os
from natsort import natsorted # pip install natsort


def add_images_to_array(folder):
    images = []
    for filename in natsorted(glob.glob(f"{folder}/*.jpg")):
        images.append(filename)

    return images


def add_images_in_excels(images, worksheet_, anchor):
    for index, image in enumerate(images):
        worksheet_.add_image(Image(image), anchor=anchor + str(index + 1))


def add_image_names_in_excel(folder):
    # titles list
    titles = []
    for title in natsorted(glob.glob(f"{folder}/*.jpg")):
        titles.append(os.path.basename(title))

    # insert titles
    for index, title in enumerate(titles):
        worksheet.cell(row=index + 1, column=3, value=title)


images_city = add_images_to_array("city_img")
grayscale_images_city = add_images_to_array("grayscale_city_img")

# create a workbook and grab active worksheet
workbook = Workbook()
worksheet = workbook.active

# resize cells
for row in range(1, 3):
    for col in range(1, 3):
        worksheet.row_dimensions[row].height = 170
        col_letter = get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = 50

# add images to excel
add_images_in_excels(images_city, worksheet, "A")
add_images_in_excels(grayscale_images_city, worksheet, "B")

# add columns with image names
add_image_names_in_excel("city_img")

# save workbook
workbook.save('image_report.xlsx')
